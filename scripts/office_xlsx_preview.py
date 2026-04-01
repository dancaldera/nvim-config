#!/usr/bin/env python3

import csv
import importlib.util
import pathlib
import re
import sys
import zipfile
from xml.etree import ElementTree as ET


MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
PKG_REL_NS = "http://schemas.openxmlformats.org/package/2006/relationships"


def qname(namespace, name):
    return f"{{{namespace}}}{name}"


def normalize_value(value):
    if value is None:
        return ""
    text = str(value)
    return re.sub(r"\s+", " ", text).strip()


def column_index(cell_ref):
    match = re.match(r"([A-Z]+)", cell_ref or "")
    if not match:
        return 0

    value = 0
    for char in match.group(1):
        value = (value * 26) + (ord(char) - 64)
    return value - 1


def joined_text(element):
    return "".join(element.itertext()) if element is not None else ""


def load_shared_strings(zf):
    try:
        root = ET.parse(zf.open("xl/sharedStrings.xml")).getroot()
    except KeyError:
        return []

    strings = []
    for item in root.findall(qname(MAIN_NS, "si")):
        strings.append(normalize_value(joined_text(item)))
    return strings


def resolve_sheet_target(zf):
    workbook = ET.parse(zf.open("xl/workbook.xml")).getroot()
    rels = ET.parse(zf.open("xl/_rels/workbook.xml.rels")).getroot()

    rel_map = {}
    for rel in rels.findall(qname(PKG_REL_NS, "Relationship")):
        rel_map[rel.get("Id")] = rel.get("Target", "")

    sheets_parent = workbook.find(qname(MAIN_NS, "sheets"))
    if sheets_parent is None:
        raise RuntimeError("Workbook does not contain any sheets.")

    sheets = sheets_parent.findall(qname(MAIN_NS, "sheet"))
    if not sheets:
        raise RuntimeError("Workbook does not contain any sheets.")

    active_index = 0
    view = workbook.find(qname(MAIN_NS, "bookViews"))
    if view is not None:
        workbook_view = view.find(qname(MAIN_NS, "workbookView"))
        if workbook_view is not None and workbook_view.get("activeTab"):
            active_index = int(workbook_view.get("activeTab", "0"))

    selected = sheets[min(active_index, len(sheets) - 1)]
    relation_id = selected.get(qname(REL_NS, "id"))
    target = rel_map.get(relation_id)
    if not target:
        raise RuntimeError("Unable to resolve worksheet path.")

    if not target.startswith("xl/"):
        target = "xl/" + target.lstrip("/")

    return selected.get("name", "Sheet1"), target


def read_sheet_with_stdlib(zf):
    shared_strings = load_shared_strings(zf)
    sheet_name, target = resolve_sheet_target(zf)
    sheet_root = ET.parse(zf.open(target)).getroot()

    rows = []
    sheet_data = sheet_root.find(qname(MAIN_NS, "sheetData"))
    if sheet_data is None:
        return sheet_name, rows

    for row in sheet_data.findall(qname(MAIN_NS, "row")):
        values = {}
        max_index = -1
        for cell in row.findall(qname(MAIN_NS, "c")):
            index = column_index(cell.get("r", ""))
            cell_type = cell.get("t")
            value = ""

            if cell_type == "inlineStr":
                value = joined_text(cell.find(qname(MAIN_NS, "is")))
            else:
                raw_value = cell.find(qname(MAIN_NS, "v"))
                if raw_value is not None and raw_value.text is not None:
                    if cell_type == "s":
                        shared_index = int(raw_value.text)
                        value = shared_strings[shared_index] if shared_index < len(shared_strings) else raw_value.text
                    elif cell_type == "b":
                        value = "TRUE" if raw_value.text == "1" else "FALSE"
                    else:
                        value = raw_value.text

            value = normalize_value(value)
            values[index] = value
            max_index = max(max_index, index)

        if max_index < 0:
            rows.append([""])
            continue

        rows.append([values.get(i, "") for i in range(max_index + 1)])

    return sheet_name, rows


def read_sheet_with_openpyxl(path):
    from openpyxl import load_workbook

    workbook = load_workbook(path, read_only=True, data_only=True)
    worksheet = workbook[workbook.active.title]
    rows = []

    for row in worksheet.iter_rows(values_only=True):
        rows.append([normalize_value(value) for value in row])

    return worksheet.title, rows


def choose_backend(path):
    if importlib.util.find_spec("openpyxl"):
        sheet_name, rows = read_sheet_with_openpyxl(path)
        return "openpyxl", sheet_name, rows

    with zipfile.ZipFile(path) as zf:
        sheet_name, rows = read_sheet_with_stdlib(zf)
    return "stdlib", sheet_name, rows


def emit_preview(path):
    backend, sheet_name, rows = choose_backend(path)
    print(f"backend={backend}", file=sys.stderr)
    print(f"sheet={sheet_name}", file=sys.stderr)

    print(f"# workbook: {path.name}")
    print(f"# sheet: {sheet_name}")

    writer = csv.writer(sys.stdout, delimiter="\t", lineterminator="\n")
    if not rows:
        writer.writerow([""])
        return

    for row in rows:
        trimmed = list(row)
        while trimmed and trimmed[-1] == "":
            trimmed.pop()
        writer.writerow(trimmed or [""])


def main():
    if len(sys.argv) != 2:
        print("usage: office_xlsx_preview.py <path>", file=sys.stderr)
        return 1

    path = pathlib.Path(sys.argv[1]).expanduser()
    if not path.exists():
        print(f"File not found: {path}", file=sys.stderr)
        return 1

    try:
        emit_preview(path)
    except Exception as exc:  # pragma: no cover - CLI fallback
        print(str(exc), file=sys.stderr)
        return 1

    return 0


if __name__ == "__main__":
    raise SystemExit(main())
